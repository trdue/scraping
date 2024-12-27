import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from datetime import datetime

def write_to_excel(file_name, df):
    """
    概要：DataFrameの内容をExcelの所定のフォーマットで出力します。
    引数：
    file_name(str):出力Excelのファイルパス
    df(DataFrame):単語とその出現回数が格納されたDataFrameを受ける。Column=["単語", "回数"]
    戻り値：(is_success: bool, message: str) 
            is_success: 処理が成功した場合は1(True)、失敗した場合は0(False)
            message: 処理結果を説明する
    使用例：
    success, message = write_to_excel(file_name, df)
    """

    #処理結果メッセージ
    message ="成功"
    #成否フラグ
    is_success = 0

    try:
        # 既存のExcelファイルを読み込む
        wb = load_workbook(file_name)
        # ブック内のアクティブなシートを取得する
        ws = wb.active

        # 今日の日付を取得
        today = datetime.now().strftime('%Y/%m/%d %H:%M')

        # dfを転置
        # 縦表示を横表示に変更する
        df_transposed = df.T

        # Excel内の書き込み開始位置を特定する
        # A列を下って空白の行を探す
        empty_row = 1  # 初めは2行目から開始
        while ws.cell(row=empty_row, column=1).value is not None:
            empty_row += 1

        # 現時点の日付をセルに書きこむ   
        ws.cell(row=empty_row, column=1, value=today)

        # DataFrameの内容をエクセルのセルに挿入:
        # まずデータを行毎にループして、その後列ごとに挿入します
        for row_num in range(len(df_transposed)):
            for col_num, value in enumerate(df_transposed.iloc[row_num,:], start=1):
                # 1行目分のデータのインデックス番号1番からデータの取り出しのループをする
                # セル毎に、順々に書き込んでいきます
                ws.cell(row=empty_row + row_num + 1, column=col_num, value=value)

        # Excelファイルに保存
        wb.save(file_name)

        is_success = 1
        return is_success, message
    except PermissionError:
            is_success = 0
            return is_success, "ファイルが開かれています。閉じてから再実行してください。"
    except Exception as e:
        is_success = 0
        message = f"エラーが発生しました: {str(e)}"
        return is_success, message

if __name__=="__main__":
    file_name = "yahoo_news.xlsx"
    dic = {
        "単語": ["yahoo","news","アクセス","ランキング",],
        "回数": ["10","30","20","1",],
    }
    df = pd.DataFrame(dic)

    success, message = write_to_excel(file_name, df)
    if success:
        print(f"成功: {message}")
    else:
        print(f"失敗: {message}")
